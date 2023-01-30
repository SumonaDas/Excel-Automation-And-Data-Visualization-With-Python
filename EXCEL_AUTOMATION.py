from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart , Reference 


ItemPrice =                  {
                                          "Chocolate": {
                                                "quantity_in_gm"        : 65,
                                                 "currentprice_in_rupee" : 78,
                                                
                                                      },
                                          
                                          "Suger": {
                                                "quantity_in_gm"        : 55,
                                                "currentprice_in_rupee" : 72
                                                
                                                      },
                                          
                                          "Meat": {
                                                "quantity_in_gm"        : 100,
                                                "currentprice_in_rupee" : 45
                                                
                                                      },
                                          
                                          "Cupcake": {
                                                "quantity_in_gm"        : 30,
                                                "currentprice_in_rupee" : 25
                                                
                                                      }
                  }


## create new workbook
wb = Workbook("D:\PPROJECTS\EXCELWITHXLSX\Data.xlsx")

## Save workbook
wb.save("D:\PPROJECTS\EXCELWITHXLSX\Data.xlsx")

## Load Existing Workbook
wb = load_workbook(filename  = "D:\PPROJECTS\EXCELWITHXLSX\Data.xlsx")


## Worksheet Object Point To An Existing Worksheet
ws_data = wb.active
ws_data.title = "Mydatasheet"


## Load Data From Dictionary
headings = ["Product"] + list(ItemPrice["Chocolate"].keys())
ws_data.append(headings)
for product in ItemPrice:
      product_details = list(ItemPrice[product].values())
      ws_data.append([product] + product_details)


## Heading  Font
for col in ws_data.iter_cols(min_col=1,min_row=1 ,max_col=1, max_row=len(ItemPrice)+1):
    for c_cell in col:
        c_cell.font = Font(bold=True)

for row in ws_data.iter_rows(min_col=1, min_row=1, max_col=len(list(ItemPrice['Chocolate'].keys()))+1, max_row=1):
    for r_cell in row :
        r_cell.font = Font(bold=True)



## Price Change Mapped To New Column
new_col=ws_data.cell(row=1, column=len(list(ItemPrice["Chocolate"].keys()))+1+1)
new_col.value ="new_price_in_rupee"
new_col.font = Font(bold=True)
for row_no in range(2,len(ItemPrice)+1+1):
      ws_data.cell(row=row_no, column=len(list(ItemPrice["Chocolate"].keys()))+1+1).value = ws_data.cell(row=row_no, column=len(list(ItemPrice["Chocolate"].keys()))+1).value*1.2



## Add Chart To New "MychartSheet" Sheet
ws_chart = wb.create_sheet("MychartSheet")
Lc = LineChart()
Lc.title = "Price Change Line Chart"
Lc.y_axis.title = "Price"
Lc.x_axis.title = "Product"


data = Reference(ws_data, min_col=3, min_row=1, max_col=len(list(ItemPrice["Chocolate"].keys()))+1+1, max_row=len(ItemPrice)+1)
Lc.add_data(data, titles_from_data=True)

ws_chart.add_chart(Lc, "A3")

## Save The Changes In File
wb.save("D:\PPROJECTS\EXCELWITHXLSX\Data.xlsx")

