from openpyxl import Workbook,load_workbook
#from openpyxl.workbook import Workbook
from openpyxl.styles import Font#,Border
import datetime as Dt
from openpyxl.chart import LineChart ,Reference 




ItemPrice =                  {
                                          "Chocolate": {
                                                "quantity_in_gm"        : 65,
                                                 "currentprice_in_rupee" : 78,
                                                
                                                      },
                                          
                                          "Suger": {
                                                "quantity_in_gm"        : 55,
                                                "currentprice_in_rupee" : 72
                                                
                                                      },
                                          
                                          "Meat": {
                                                "quantity_in_gm"        : 100,
                                                "currentprice_in_rupee" : 45
                                                
                                                      },
                                          
                                          "Cupcake": {
                                                "quantity_in_gm"        : 30,
                                                "currentprice_in_rupee" : 25
                                                
                                                      }
                  }

#Product	quantity_in_gm	currentprice_in_rupee	new_price
#Chocolate	65	78	93.6
#Suger	55	72	86.4
#Meat	      100	45	54
#Cupcake	30	25	30
    

## create new workbook
#wb = Workbook("D:\PPROJECTS\EXCELWITHXLSX\Data.xlsx") # print(wb.sheetnames) # WriteOnlyWorksheet
wb = load_workbook(filename  = "D:\PPROJECTS\EXCELWITHXLSX\Data.xlsx")


## create  worksheet for data load
# When a worksheet is created in memory, it contains no cells. They are created when first accessed
# Data can be assigned directly to cells # ws["A1"] = 42 # cell_data_for_b2 = ws.cell(row=4, column=2, value=10)
# cell_range = ws["A1":"C2"]
#colC = ws["C"]  col_range = ws["C:D"]
#row10 = ws[10]  row_range = ws[5:10]  
# Rows can also be appended # ws1.append([1, 2, 3]) # after active row
#for col in ws.iter_cols(min_row=1, max_col=3, max_row=2): for cell in col: print(cell.values)
#for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):for cell in row:print(cell.values)
#for row in data: ws.append(row)

## workbsheet object
#ws_data = wb.create_sheet("Mydatasheet")  # WriteOnlyWorksheet # for sheet in wb : print(sheet.title) # ws1.title = "New Title" # ws1 = wb ["New Title"]

ws_data = wb.active
ws_data.title = "Mydatasheet"


## Data Read
headings = ["Product"] + list(ItemPrice["Chocolate"].keys())
ws_data.append(headings)
for product in ItemPrice:
      product_details = list(ItemPrice[product].values())
      ws_data.append([product] + product_details)


##Bold Heading  
for col in ws_data.iter_cols(min_row=1, max_col=1, max_row=len(ItemPrice)+1):
    for c_cell in col:
        c_cell.font = Font(bold=True)

for row in ws_data.iter_rows(min_row=1, max_col=len(list(ItemPrice['Chocolate'].keys()))+1, max_row=1):
    for r_cell in row :
        r_cell.font = Font(bold=True)


''' # for WriteOnlyWorksheet have to use it
cols = ws_data['A1':'G1']
rows = ws_data['A1':'A5']




for col in cols:
      for c_cell in col:
            c_cell.font = Font(bold=True)

for row in rows:
      for r_cell in row :
            r_cell.font = Font(bold=True)
'''

## Price Change Citeria
Current_Year = Dt.date.today().year # datetime.datetime.today().year
if Current_Year == 2023:
      ws_data.cell(row=1, column=len(list(ItemPrice["Chocolate"].keys()))+1+1).value ="new_price"
      ws_data.cell(row=1, column=len(list(ItemPrice["Chocolate"].keys()))+1+1).font = Font(bold=True)
      for row_no in range(2,len(ItemPrice)+1+1):
            ws_data.cell(row=row_no, column=len(list(ItemPrice["Chocolate"].keys()))+1+1).value = ws_data.cell(row=row_no, column=len(list(ItemPrice["Chocolate"].keys()))+1).value*1.2



## Add Chart To "Mychartsheet" Sheet
ws_chart = wb.create_sheet("Mychartsheet")


Lc = LineChart()
Lc.title = "Price Change Line Chart"
#Lc.style = 10
Lc.y_axis.title = "Price"
Lc.x_axis.title = "Product"


data = Reference(ws_data, min_col=3, min_row=1, max_col=len(list(ItemPrice["Chocolate"].keys()))+1+1, max_row=len(ItemPrice)+1)
Lc.add_data(data, titles_from_data=True)

ws_chart.add_chart(Lc, "A5")




## Save the file
wb.save("D:\PPROJECTS\EXCELWITHXLSX\Data.xlsx")
wb.close()

########################################################################################################################################################################################################

"""

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ["Name"] + list(data["Joe"].keys())
ws.append(headings)

for person in data:
      grades = list(data[person].values())
      ws.append([person] + grades)




for col in range(2, len(data["Joe"]) + 2):
      char = get_column_letter(col)
      ws[char + "7"] = f"=SUM({char + "2"}:{char + "6"})/{len(data)}"

for col in range(1, 6):
      ws[get_column_letter(col) + "1"].font = Font(bold=True, color="0099CCFF")

wb.save("NewGrades.xlsx")


###################
pyvot,chart,formula

"""